import os
import io
import base64
import logging
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from fastapi.responses import Response
from pydantic import BaseModel, field_validator
from api.dependencies import get_current_user
from api.sage import chat
from api import security_log, rate_limit

logger = logging.getLogger("vaultic.sage")
router = APIRouter(prefix="/api/sage", tags=["sage"])

MAX_FILE_BYTES = 20 * 1024 * 1024   # 20 MB
MAX_IMAGE_BYTES = 5 * 1024 * 1024   # 5 MB (Anthropic base64 limit)

IMAGE_TYPES = {
    "image/jpeg":  ".jpg",
    "image/png":   ".png",
    "image/gif":   ".gif",
    "image/webp":  ".webp",
    "image/bmp":   ".bmp",
    "image/tiff":  ".tiff",
}

EXTENSION_MIME = {
    ".jpg": "image/jpeg", ".jpeg": "image/jpeg",
    ".png": "image/png", ".gif": "image/gif",
    ".webp": "image/webp", ".bmp": "image/bmp", ".tiff": "image/tiff",
}


class ChatRequest(BaseModel):
    message: str
    history: list[dict] = []
    attachments: list[dict] = []   # [{type, content, media_type, filename}]

    @field_validator("message")
    @classmethod
    def message_max_length(cls, v):
        if len(v) > 10_000:
            raise ValueError("Message must be under 10,000 characters")
        return v

    @field_validator("history")
    @classmethod
    def history_max_entries(cls, v):
        if len(v) > 100:
            raise ValueError("History must be under 100 messages")
        return v


class ChatResponse(BaseModel):
    response: str
    history: list[dict]


class SpeakRequest(BaseModel):
    text: str


@router.post("/chat", response_model=ChatResponse)
async def sage_chat(body: ChatRequest, _user: str = Depends(get_current_user)):
    limited, remaining = rate_limit.check_sage(_user)
    if limited:
        security_log.log_server_event(f"SAGE_RATE_LIMITED  user={_user}")
        raise HTTPException(status_code=429, detail="Sage message limit reached (60/hour). Try again later.")
    rate_limit.record_sage(_user)
    security_log.log_sage_query(_user, body.message)
    try:
        response, updated_history = chat(body.history, body.message, body.attachments)
        return ChatResponse(response=response, history=updated_history)
    except Exception as e:
        logger.error(f"Sage chat error: {e}")
        raise HTTPException(status_code=500, detail="Internal server error")


@router.post("/process-file")
async def process_file(
    file: UploadFile = File(...),
    _user: str = Depends(get_current_user),
):
    """
    Extract content from an uploaded file for Sage to read.
    Images are returned as base64 for direct Claude vision processing.
    All other types are returned as extracted text.
    """
    data = await file.read()
    if len(data) > MAX_FILE_BYTES:
        raise HTTPException(status_code=413, detail=f"File too large (max 20 MB)")

    filename = file.filename or "file"
    ext = ("." + filename.rsplit(".", 1)[-1].lower()) if "." in filename else ""
    content_type = file.content_type or ""

    # Detect image by MIME or extension
    mime = content_type if content_type in IMAGE_TYPES else EXTENSION_MIME.get(ext, "")
    if mime:
        if len(data) > MAX_IMAGE_BYTES:
            raise HTTPException(status_code=413, detail="Image too large (max 5 MB for vision)")
        return {
            "type": "image",
            "media_type": mime,
            "content": base64.standard_b64encode(data).decode(),
            "filename": filename,
        }

    # Text extraction by file type
    try:
        text = _extract_text(data, ext, filename)
    except Exception as e:
        raise HTTPException(status_code=422, detail=f"Could not read {filename}: {e}")

    # Cap text at 100k chars to stay within context limits
    truncated = len(text) > 100_000
    return {
        "type": "text",
        "content": text[:100_000],
        "filename": filename,
        "truncated": truncated,
    }


def _extract_text(data: bytes, ext: str, filename: str) -> str:
    """Extract plain text from a file given its raw bytes and extension."""

    # PDF
    if ext == ".pdf":
        import pdfplumber
        with pdfplumber.open(io.BytesIO(data)) as pdf:
            pages = [p.extract_text() or "" for p in pdf.pages]
        return "\n\n".join(pages).strip()

    # Word (.docx)
    if ext == ".docx":
        import docx
        doc = docx.Document(io.BytesIO(data))
        return "\n".join(p.text for p in doc.paragraphs if p.text.strip())

    # Excel (.xlsx)
    if ext in (".xlsx", ".xlsm", ".xltx"):
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        parts = []
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            rows = []
            for row in ws.iter_rows(values_only=True):
                row_vals = [str(c) if c is not None else "" for c in row]
                if any(v.strip() for v in row_vals):
                    rows.append("\t".join(row_vals))
            if rows:
                parts.append(f"--- Sheet: {sheet} ---\n" + "\n".join(rows))
        return "\n\n".join(parts)

    # Legacy Excel (.xls)
    if ext == ".xls":
        try:
            import xlrd
            wb = xlrd.open_workbook(file_contents=data)
            parts = []
            for sheet in wb.sheets():
                rows = []
                for r in range(sheet.nrows):
                    row_vals = [str(sheet.cell_value(r, c)) for c in range(sheet.ncols)]
                    if any(v.strip() for v in row_vals):
                        rows.append("\t".join(row_vals))
                if rows:
                    parts.append(f"--- Sheet: {sheet.name} ---\n" + "\n".join(rows))
            return "\n\n".join(parts)
        except ImportError:
            return "Legacy .xls files require xlrd — install with: pip install xlrd"

    # YAML
    if ext in (".yaml", ".yml"):
        import yaml
        obj = yaml.safe_load(data)
        return yaml.dump(obj, default_flow_style=False)

    # XML / HTML — strip tags, return readable text
    if ext in (".xml", ".html", ".htm"):
        import re
        text = data.decode("utf-8", errors="replace")
        text = re.sub(r"<style[^>]*>.*?</style>", "", text, flags=re.DOTALL)
        text = re.sub(r"<script[^>]*>.*?</script>", "", text, flags=re.DOTALL)
        text = re.sub(r"<[^>]+>", " ", text)
        return re.sub(r"\s{3,}", "\n\n", text).strip()

    # Everything else: decode as UTF-8 text
    # Covers: .txt, .md, .csv, .json, .js, .ts, .py, .sql, .log, .ini, .toml, etc.
    return data.decode("utf-8", errors="replace")


@router.post("/transcribe")
async def transcribe_audio(
    file: UploadFile = File(...),
    _user: str = Depends(get_current_user),
):
    """
    Transcribe speech to text using OpenAI Whisper.
    Accepts webm/ogg/mp4/wav/m4a audio from the browser's MediaRecorder.
    Significantly more accurate than the browser's Web Speech API, especially
    for financial terms, numbers, and proper nouns.
    """
    api_key = os.environ.get("OPENAI_API_KEY", "")
    if not api_key:
        raise HTTPException(status_code=503, detail="OPENAI_API_KEY not configured")

    data = await file.read()
    if len(data) > 25 * 1024 * 1024:
        raise HTTPException(status_code=413, detail="Audio too large (max 25 MB)")
    if len(data) < 100:
        raise HTTPException(status_code=400, detail="Audio clip too short")

    try:
        import openai
        client = openai.AsyncOpenAI(api_key=api_key)
        filename = file.filename or "audio.webm"
        result = await client.audio.transcriptions.create(
            model="whisper-1",
            file=(filename, io.BytesIO(data), file.content_type or "audio/webm"),
            language="en",
        )
        return {"text": result.text}
    except Exception as e:
        logger.error(f"Whisper transcription error: {e}")
        raise HTTPException(status_code=500, detail="Internal server error")


@router.post("/speak")
async def sage_speak(body: SpeakRequest, _user: str = Depends(get_current_user)):
    """
    Convert text to speech using OpenAI TTS (async). Returns the full MP3.
    Voice "fable" — warm, expressive. Valid tts-1 voices: nova, shimmer, echo,
    onyx, fable, alloy, ash, sage, coral.
    """
    api_key = os.environ.get("OPENAI_API_KEY", "")
    if not api_key:
        raise HTTPException(status_code=503, detail="OPENAI_API_KEY not configured")

    try:
        import openai
        client = openai.AsyncOpenAI(api_key=api_key)
        response = await client.audio.speech.create(
            model="tts-1",
            voice="fable",
            input=body.text[:4096],
            response_format="mp3",
            speed=1.2,  # 1.2x feels natural for a financial advisor; 1.0 is too slow
        )
        return Response(response.content, media_type="audio/mpeg")
    except Exception as e:
        logger.error(f"TTS error: {e}")
        raise HTTPException(status_code=500, detail="Internal server error")
